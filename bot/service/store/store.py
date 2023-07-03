# Excel
from openpyxl import load_workbook, worksheet, workbook
from pydantic import SecretStr

# Google sheets
import httplib2
import apiclient.discovery
from oauth2client.service_account import ServiceAccountCredentials

from core.entities import ProfileLink


class Excel_interactions:
    path: str

    def __init__(self, excel_filename) -> None:
        self.path = excel_filename

    def read_xlsx(self) -> worksheet:
        excel = load_workbook(filename=self.path)
        return excel['Лист1']
    
    def write_xlsx(self) -> workbook.Workbook:
        return load_workbook(filename=self.path)

    def extract(self, day: int = None) -> list[list]:
        timetable_cells = list()

        timetable_xlsx = self.read_xlsx()
        for row in timetable_xlsx.iter_rows(min_row=2, min_col=2, max_col=7, max_row=5):
            row_as_lst = list()
            for n, cell in enumerate(row):
                if day is None or (day == n and day is not None):
                    if "<a" in cell.value:

                        row_as_lst.append(
                            ProfileLink(
                            **{
                                'id': cell.value.split('id=')[1].split('"')[0],
                                'fullname': cell.value.split('</a>')[0].split('">')[1]
                            }))
                    else:
                        row_as_lst.append(cell.value)
            timetable_cells.append(row_as_lst)

        return timetable_cells

    def put(self, data, position: tuple[str, str]) -> worksheet:
        timetable_wb = self.write_xlsx()
        timetable_sheet = timetable_wb['Лист1']

        timetable_sheet[position[0] + position[1]] = data
        timetable_wb.save(self.path)
        return timetable_sheet


class GoogleSheet_interactions:
    spreadsheetId: str
    service: apiclient
    credentials: SecretStr

    def __init__(self, CREDENTIALS_FILE: str, spreadsheetId: str):
        self.spreadsheetId = spreadsheetId

        # Auth
        self.credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive'])
        httpAuth = self.credentials.authorize(httplib2.Http())
        self.service = apiclient.discovery.build('sheets', 'v4', http = httpAuth)

    def __create_table__(self):
        spreadsheet = self.service.spreadsheets().create(body = {'properties': {'title': 'Первый тестовый документ', 'locale': 'ru_RU'}, 
                                                            'sheets': [{'properties': {'sheetType': 'GRID',
                                                            'sheetId': 0,
                                                            'title': 'Лист номер один',
                                                            'gridProperties': {'rowCount': 100, 'columnCount': 15}}}]
                                                            }).execute()
        spreadsheetId = spreadsheet['spreadsheetId'] # Don't forget save spreadsheetID!
    
    def __create_sheet(self):
        # Добавление листа
        results = self.service.spreadsheets().batchUpdate(
            spreadsheetId = self.spreadsheetId,
            body = 
        {
        "requests": [
            {
            "addSheet": {
                "properties": {
                "title": "Another_list",
                "gridProperties": {
                    "rowCount": 20,
                    "columnCount": 12
                }
                }
            }
            }
        ]
        }).execute()
    
    def __get_list_of_sheets(self):
        # Получаем список листов, их Id и название
        spreadsheet = self.service.spreadsheets().get(spreadsheetId = spreadsheetId).execute()
        sheetList = spreadsheet.get('sheets')
        # for sheet in sheetList:
        #     print(sheet['properties']['sheetId'], sheet['properties']['title'])
            
        sheetId = sheetList[0]['properties']['sheetId']
        return sheetList


    def grant_permission(self, email: str):
        driveService = apiclient.discovery.build('drive', 'v3', http = httpAuth)
        driveService.permissions().create(fileId = self.spreadsheetId, body = {
            'type': 'user',
            'role': 'writer',
            'emailAddress': email
            }, fields = 'id').execute()
   
    def extract(self, day: int = None):
        if day is None:
            ranges = ["Лист 1!B2:G5"]
        else:
            ranges = [f"Лист 1!{chr(day+66)}2:{chr(day+66)}5"]

        results = self.service.spreadsheets().values().batchGet(spreadsheetId = self.spreadsheetId,
                                            ranges = ranges,
                                            dateTimeRenderOption = 'FORMATTED_STRING').execute()
        sheet_values = results['valueRanges'][0]['values']

        timetable_cells = list()
        for row in sheet_values:
            row_as_lst = list()
            for n, cell in enumerate(row):
                if "<a" in cell:

                    row_as_lst.append(
                        ProfileLink(
                        **{
                            'id': cell.split('id=')[1].split('"')[0],
                            'fullname': cell.split('</a>')[0].split('">')[1]
                        }))
                else:
                    row_as_lst.append(cell)
            timetable_cells.append(row_as_lst)

        return timetable_cells


    def put(self, data: str, position: tuple[str, str]):
        results = self.service.spreadsheets().values().batchUpdate(spreadsheetId = self.spreadsheetId, body = {
        "valueInputOption": "USER_ENTERED", # Данные воспринимаются, как вводимые пользователем (считается значение формул)
        "data": [
            {"range": f"Лист 1!{position[0]}{position[1]}",
            "majorDimension": "ROWS",     # Сначала заполнять строки, затем столбцы
            "values": [
                        [data, ]
                    ]}
        ]
        }).execute()





class DB_interactions:
    pass
    def extracting_from_excel():

        pass


    def put_to_excel():
        pass

    def put_to_db():
        pass
