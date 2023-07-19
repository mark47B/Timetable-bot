from .store import Store_interaction
from openpyxl import load_workbook, worksheet, workbook
from core.entities import AVAILABLE_DAYS, AVAILABLE_TIME, parse_cell_to_ProfileLink, ProfileLink


class Excel_interactions(Store_interaction):
    path: str

    def __init__(self, excel_filename) -> None:
        self.path = excel_filename

    def read_xlsx(self) -> worksheet:
        excel = load_workbook(filename=self.path)
        return excel['Лист1']

    def write_xlsx(self) -> workbook.Workbook:
        return load_workbook(filename=self.path)

    def extract(self, day: int = None) -> list[list]:
        timetable_cells = list()

        timetable_xlsx = self.read_xlsx()
        for row in timetable_xlsx.iter_rows(min_row=2, min_col=2, max_col=7, max_row=5):
            row_as_lst = list()
            for n, cell in enumerate(row):
                if day is None or (day == n and day is not None):
                    if "username='@" in cell.value:
                        row_as_lst.append(parse_cell_to_ProfileLink(cell.value))
                    else:
                        row_as_lst.append(cell.value)
            timetable_cells.append(row_as_lst)

        return timetable_cells

    def put(self, data: ProfileLink, position: tuple[str, str]) -> worksheet:
        timetable_wb = self.write_xlsx()
        timetable_sheet = timetable_wb['Лист1']

        timetable_sheet[position[0] + position[1]] = data.excel_repr()
        timetable_wb.save(self.path)
        return timetable_sheet

    def get_free_slots(self) -> dict:
        whole_timetable = self.extract()

        free_slots = dict()
        for number_row, row in enumerate(whole_timetable):
            for number_column, cell in enumerate(row):
                if cell == 'Not Reserved':
                    if AVAILABLE_DAYS[number_column] in free_slots.keys():
                        free_slots[AVAILABLE_DAYS[number_column]] += [AVAILABLE_TIME[number_row] ]
                    else:
                        free_slots[AVAILABLE_DAYS[number_column]] = [AVAILABLE_TIME[number_row] ]
        return free_slots
